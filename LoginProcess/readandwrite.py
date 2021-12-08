import openpyxl
from selenium import webdriver
import time

Execution = "Yes"
UseCase = "durga"
path = "C:/Users/dvandhar/PycharmProjects/pythonProject/execute.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

for r in range(1, 20):
    print("first loop" + str(r))
    if sheet.cell(r, 1).value == None:
        print("loop is empty")
        break
    else:
        if sheet.cell(r, 1).value in UseCase and sheet.cell(r, 2).value in "Yes":
            sheet.cell(row=r, column=3).value = "Correct"
        elif sheet.cell(r, 1).value in UseCase and sheet.cell(r, 2).value in "No":
            sheet.cell(row=r, column=3).value = "InCorrect"
        elif sheet.cell(r, 1).value not in UseCase and sheet.cell(r, 2).value in "Yes" or sheet.cell(r, 2).value in "No" :
            sheet.cell(row=r, column=3).value = "InCorrect"
print()
if Execution in "Yes":
    print("Execution of the process " + Execution)
    driver = webdriver.Chrome(executable_path="C:/Users/dvandhar/PycharmProjects/pythonProject/Chrome/chromedriver.exe")
    driver.implicitly_wait(5)
    driver.maximize_window()
    driver.get("https://www.google.co.in/")
else:
    print(Execution)
wb.save("C:/Users/dvandhar/PycharmProjects/pythonProject/execute.xlsx")
driver.close()
