from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl

#To Open chrome browser
driver=webdriver.Chrome(executable_path="C:/Users/dvandhar/PycharmProjects/pythonProject/Chrome/chromedriver.exe")

#Working with excel sheet
path = "C:/Users/dvandhar/Downloads/testScenarios.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
#To check maximum no. of rows
rows = sheet.max_row
print(rows)
#To check maximum no. of columns
cols = sheet.max_column
print(cols)
for r in range(2,6):
    print("row is " +str(r))
    driver.get("https://login.dovico.com/#Login")
    time.sleep(2)
    # Company
    value1 = sheet.cell(row=r, column=1).value
    print(value1)
    driver.find_element_by_xpath("//*[@id="loginControls"]/div[2]/input[2]").send_keys(value1)
    time.sleep(2)
    #Username
    value1 = sheet.cell(row=r, column=1).value
    print(value1)
    driver.find_element_by_xpath("//*[@id="loginControls"]/div[3]/input").send_keys(value1)
    time.sleep(2)
    #Password
    value2 = sheet.cell(row=r, column=2).value
    print(value2)
    driver.find_element_by_xpath("//*[@id="loginControls"]/div[4]/input").send_keys(value2)
    time.sleep(2)
    #Login Button
    driver.find_element_by_xpath("/html/body/div[1]/div/div[1]/div/form[4]/div[3]/div/input[2]").click()
    time.sleep(4)

    if driver.title == "Appian Designer":
        print("Test Case is Passed")
        print(driver.title)
        credentials=sheet.cell(row=r, column=3).value = "Correct"
        print(credentials)
        sheet.cell(row=r, column=4).value = "Test Case is Passed"
    else:
        print("Test Case is Failed")
        print(driver.title)
        credentials=sheet.cell(row=r, column=3).value = "Wrong"
        sheet.cell(row=r, column=4).value = "Test Case is Failed"
        print(credentials)
    wb.save("PathOfExcelSheet")
driver.close()