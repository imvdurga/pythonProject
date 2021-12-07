import openpyxl
from selenium import webdriver
import time

Execution = "Yes"
UseCase = "Execution durga"
path = "C:/Users/dvandhar/Downloads/execute.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

for row in range(1,10):
    print("first loop"+str(row))
    if sheet.cell(row,1).value == None:
        print("loop is empty")
        break
    else:
        if sheet.cell(row,1).value in UseCase:
            print("Use Case Is Correct " +sheet.cell(row,1).value)
            if sheet.cell(row,2).value in "No":
                print("Can't perform the Execution because its failing with NO " +sheet.cell(row,2).value)
                Execution = "No"
            elif sheet.cell(row,2).value in "Yes":
                print("Can perform the Execution because it satisfies with YES " +sheet.cell(row,2).value)
                Execution = "Yes"
print()
if Execution in "Yes":
    print("Execution of the process "+Execution)
    driver = webdriver.Chrome(executable_path="C:/Users/dvandhar/PycharmProjects/pythonProject/Chrome/chromedriver.exe")
    driver.implicitly_wait(5)
    driver.maximize_window()
    driver.get("https://www.facebook.com/")
    driver.find_element_by_xpath(
        "//div[@class='_8iep _8icy _9ahz _9ah-']/div[1]/form/div[1]/div[1]/input").send_keys(9963636739)
    driver.find_element_by_xpath(
        "//div[@class='_8iep _8icy _9ahz _9ah-']/div[1]/form/div[1]/div[2]/div/input").send_keys("imvdurga@13Prasad")
    driver.find_element_by_xpath(
        "//div[@class='_8iep _8icy _9ahz _9ah-']/div[1]/form/div[2]/button").click()
    time.sleep(5)
else:
    print(Execution)
driver.close()
